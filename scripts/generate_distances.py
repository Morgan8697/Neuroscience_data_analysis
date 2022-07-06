from math import dist
import openpyxl
from cell import Cell, Family
import pandas as pd 

PATH_DATA = "Neuroscience_data_analysis\data\PNN_Microglia_distance.xlsx"
SHEET = 's512 B2 centered (1) detection '
MAX_DISTANCE = 50

wb = openpyxl.load_workbook(PATH_DATA)

def generate_distances(worksheet_name, sorted_cells, type):
    sheet = wb[worksheet_name]
    wb.active = sheet
    micro_list = sorted_cells[type]['microglia_cells_list']
    pnn_list = sorted_cells[type]['pnn_cells_list']
    sheet.delete_cols(1,200)
    sheet.delete_rows(1,200)
    sheet.cell(row=1, column=1).value="Microglia"
    sheet.cell(row=1, column=2).value="PNN"
    sheet.cell(row=1, column=3).value="Distance"
    row = 2
    distances = []
    # Calulating distances
    for microglia_cell in micro_list:
        for pnn_cell in pnn_list:
            point_microglia = [microglia_cell.position_x, microglia_cell.position_y]
            point_pnn = [pnn_cell.position_x, pnn_cell.position_y]
            distance = dist(point_microglia, point_pnn)
            if distance <= MAX_DISTANCE:
                sheet.cell(row=row, column=1).value=microglia_cell.id
                sheet.cell(row=row, column=2).value=pnn_cell.id
                sheet.cell(row=row, column=3).value=distance
                distances.append(distance)
                row += 1

    return pd.Series(data=distances)
            
sheet1 = wb[SHEET]
wb.active = sheet1
row_count = sheet1.max_row
col_count = sheet1.max_column

sorted_cells = {
    'L1_2' : {
        'microglia_cells_list' : [],
        'pnn_cells_list' : []
        },
    'L3' : {
        'microglia_cells_list' : [],
        'pnn_cells_list' : []
        },
    'L4' : {
        'microglia_cells_list' : [],
        'pnn_cells_list' : []
        },
    'L5' : {
        'microglia_cells_list' : [],
        'pnn_cells_list' : []
        },
    'L6' : {
        'microglia_cells_list' : [],
        'pnn_cells_list' : []
        }
}
# Initializing Cells
for i in range(2, row_count + 1):
    curr_name = sheet1.cell(row=i, column=1).value
    curr_type = curr_name.split(' - ')[1]
    curr_family = Family.MICROGLIA
    if "PNN" in curr_name:
        curr_family = Family.PNN
    curr_x = sheet1.cell(row=i, column=2).value
    curr_y = sheet1.cell(row=i, column=3).value
    if curr_family == Family.MICROGLIA:
        curr_id = "Microglia #" + str(i)
        curr_cell = Cell(curr_id, curr_x, curr_y, curr_family, curr_type)
        sorted_cells[curr_type]['microglia_cells_list'].append(curr_cell)
    else:
        curr_id = "PNN #" + str(i)
        curr_cell = Cell(curr_id, curr_x, curr_y, curr_family, curr_type)
        sorted_cells[curr_type]['pnn_cells_list'].append(curr_cell)

# Sort seriess
    
L1_series = generate_distances('Distances L1_2', sorted_cells, 'L1_2')
L3_series = generate_distances('Distances L3', sorted_cells, 'L3')
L4_series = generate_distances('Distances L4', sorted_cells, 'L4')
L5_series = generate_distances('Distances L5', sorted_cells, 'L5')
L6_series = generate_distances('Distances L6', sorted_cells, 'L6')

# Data analysis
print("----------\nL1_2\nAverage: " + str(L1_series.mean()) + "\nMedian: " + str(L1_series.median()) + "\nStandard deviation: " + str(L1_series.std()))
print("----------\nL3\nAverage: " + str(L3_series.mean()) + "\nMedian: " + str(L3_series.median()) + "\nStandard deviation: " + str(L3_series.std()))
print("----------\nL4\nAverage: " + str(L4_series.mean()) + "\nMedian: " + str(L4_series.median()) + "\nStandard deviation: " + str(L4_series.std()))
print("----------\nL5\nAverage: " + str(L5_series.mean()) + "\nMedian: " + str(L5_series.median()) + "\nStandard deviation: " + str(L5_series.std()))
print("----------\nL6\nAverage: " + str(L6_series.mean()) + "\nMedian: " + str(L6_series.median()) + "\nStandard deviation: " + str(L6_series.std()))

wb.save(PATH_DATA)

    